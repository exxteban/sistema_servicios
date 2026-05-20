"""
Rutas de gestión de clientes
"""
import re
from io import BytesIO
from datetime import date, datetime
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, make_response
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_DISPONIBLE = True
except ModuleNotFoundError:
    Workbook = None
    Alignment = Border = Font = PatternFill = Side = None
    get_column_letter = None
    OPENPYXL_DISPONIBLE = False
from app import db
from app.models import Cliente, ClienteObservacion, Reparacion, Venta, DetalleReparacion
from app.services.clientes_fidelizacion import obtener_resumen_beneficios_cliente, sincronizar_beneficios_vencidos, beneficio_resumen_snapshot
from app.services.clientes_fidelizacion_sincronizacion import sincronizar_compras_fidelizacion_pendientes
from app.utils.permisos import validar_autorizacion
from app.utils.helpers import local_strftime
from app.utils.phone_utils import normalizar_telefono

clientes_bp = Blueprint('clientes', __name__)


def _parsear_nivel_estrellas(raw):
    try:
        nivel = int(raw or 3)
    except (TypeError, ValueError):
        nivel = 3
    return max(1, min(5, nivel))


def _a_float_seguro(valor, default=0.0):
    if valor is None:
        return default
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto:
        return default
    texto = texto.replace('.', '').replace(',', '.') if ',' in texto and '.' in texto else texto.replace(',', '.')
    try:
        return float(texto)
    except (TypeError, ValueError):
        return default


def _estado_reparacion_display(estado):
    if not estado:
        return ''
    labels = {
        'pendiente': 'Pendiente',
        'diagnostico': 'En Diagnóstico',
        'espera_presupuesto': 'A confirmar presupuesto',
        'espera_repuesto': 'Espera repuesto',
        'espera_cliente': 'Espera cliente',
        'en_proceso': 'En proceso',
        'listo': 'Listo',
        'no_se_pudo': 'No se pudo',
        'entregado': 'Entregado',
        'cancelado': 'Cancelado',
        'antiguos': 'Antiguos',
    }
    return labels.get(estado, str(estado).replace('_', ' ').title())


def _historial_sort_key(item):
    fecha = item.get('fecha')
    if isinstance(fecha, datetime):
        return (fecha.year, fecha.month, fecha.day, fecha.hour, fecha.minute, fecha.second, fecha.microsecond)
    if isinstance(fecha, date):
        return (fecha.year, fecha.month, fecha.day, 0, 0, 0, 0)
    return (0, 0, 0, 0, 0, 0, 0)


def _construir_query_clientes(buscar, solo_top, sort_by, sort_dir, min_estrellas):
    total_compras_expr = db.select(
        db.func.count(Venta.id_venta)
    ).where(
        Venta.id_cliente == Cliente.id_cliente,
        Venta.estado == 'completada'
    ).scalar_subquery()
    total_reparaciones_expr = db.select(
        db.func.count(Reparacion.id_reparacion)
    ).where(
        Reparacion.cliente_id == Cliente.id_cliente
    ).scalar_subquery()
    total_movimientos_expr = total_compras_expr + total_reparaciones_expr

    query = Cliente.query.filter_by(activo=True)

    if buscar:
        buscar_celular = re.sub(r'\D', '', buscar)
        telefono_normalizado = db.func.replace(
            db.func.replace(
                db.func.replace(
                    db.func.replace(
                        db.func.replace(
                            db.func.replace(Cliente.telefono, ' ', ''),
                            '-', ''
                        ),
                        '(', ''
                    ),
                    ')', ''
                ),
                '+', ''
            ),
            '.', ''
        )
        condiciones_busqueda = [
            Cliente.nombre.ilike(f'%{buscar}%'),
            Cliente.ruc_ci.ilike(f'%{buscar}%'),
            Cliente.telefono.ilike(f'%{buscar}%')
        ]
        if buscar_celular:
            condiciones_busqueda.append(telefono_normalizado.ilike(f'%{buscar_celular}%'))
        query = query.filter(db.or_(*condiciones_busqueda))
    if solo_top:
        query = query.filter(Cliente.nivel_estrellas >= min_estrellas)

    columnas_ordenables = {
        'nombre': Cliente.nombre,
        'ruc_ci': Cliente.ruc_ci,
        'telefono': Cliente.telefono,
        'tipo': Cliente.tipo,
        'total_compras': total_compras_expr,
        'total_reparaciones': total_reparaciones_expr,
        'total_movimientos': total_movimientos_expr,
        'nivel': Cliente.nivel_estrellas,
    }
    if sort_by not in columnas_ordenables:
        sort_by = 'nombre'
    if sort_dir not in ('asc', 'desc'):
        sort_dir = 'asc'
    columna_orden = columnas_ordenables[sort_by]
    orden_primario = columna_orden.desc() if sort_dir == 'desc' else columna_orden.asc()
    orden_secundario = Cliente.id_cliente.desc() if sort_dir == 'desc' else Cliente.id_cliente.asc()
    return query.order_by(orden_primario, orden_secundario), sort_by, sort_dir


def _inyectar_totales_clientes(clientes):
    clientes_ids = [cliente.id_cliente for cliente in clientes]
    compras_por_cliente = {}
    reparaciones_por_cliente = {}
    if clientes_ids:
        compras_por_cliente = dict(
            db.session.query(
                Venta.id_cliente,
                db.func.count(Venta.id_venta)
            ).filter(
                Venta.id_cliente.in_(clientes_ids),
                Venta.estado == 'completada'
            ).group_by(Venta.id_cliente).all()
        )
        reparaciones_por_cliente = dict(
            db.session.query(
                Reparacion.cliente_id,
                db.func.count(Reparacion.id_reparacion)
            ).filter(
                Reparacion.cliente_id.in_(clientes_ids)
            ).group_by(Reparacion.cliente_id).all()
        )
    for cliente in clientes:
        cliente.total_compras = int(compras_por_cliente.get(cliente.id_cliente, 0) or 0)
        cliente.total_reparaciones = int(reparaciones_por_cliente.get(cliente.id_cliente, 0) or 0)
        cliente.total_movimientos = cliente.total_compras + cliente.total_reparaciones


@clientes_bp.route('/')
@login_required
def listar():
    """Lista de clientes"""
    if not current_user.tiene_permiso('ver_clientes'):
        if getattr(current_user, 'modo_demo', False):
            flash('Modo demo: esta acción está deshabilitada.', 'warning')
        else:
            flash('No tienes permisos para ver clientes.', 'danger')
        return redirect(url_for('main.dashboard'))

    page = max(request.args.get('page', 1, type=int), 1)
    per_page = request.args.get('per_page', 20, type=int)
    if per_page not in (10, 20, 50, 100):
        per_page = 20
    buscar = request.args.get('buscar', '')
    solo_top = request.args.get('solo_top', '0') == '1'
    sort_by = (request.args.get('sort_by') or 'nombre').strip().lower()
    sort_dir = (request.args.get('sort_dir') or 'asc').strip().lower()
    min_estrellas = 4
    query_ordenada, sort_by, sort_dir = _construir_query_clientes(
        buscar=buscar,
        solo_top=solo_top,
        sort_by=sort_by,
        sort_dir=sort_dir,
        min_estrellas=min_estrellas
    )
    
    total_clientes = query_ordenada.count()
    if total_clientes > 0:
        max_page = ((total_clientes - 1) // per_page) + 1
        page = min(page, max_page)
    clientes = query_ordenada.paginate(
        page=page, per_page=per_page, error_out=False
    )
    _inyectar_totales_clientes(clientes.items)
    
    return render_template(
        'clientes/listar.html',
        clientes=clientes,
        buscar=buscar,
        solo_top=solo_top,
        min_estrellas=min_estrellas,
        sort_by=sort_by,
        sort_dir=sort_dir,
        per_page=per_page
    )


@clientes_bp.route('/exportar_xlsx')
@login_required
def exportar_xlsx():
    if not current_user.tiene_permiso('ver_clientes'):
        if getattr(current_user, 'modo_demo', False):
            flash('Modo demo: esta acción está deshabilitada.', 'warning')
        else:
            flash('No tienes permisos para exportar clientes.', 'danger')
        return redirect(url_for('main.dashboard'))
    if not OPENPYXL_DISPONIBLE:
        flash('La exportación a Excel no está disponible en este servidor.', 'danger')
        return redirect(url_for('clientes.listar'))

    buscar = request.args.get('buscar', '')
    solo_top = request.args.get('solo_top', '0') == '1'
    sort_by = (request.args.get('sort_by') or 'nombre').strip().lower()
    sort_dir = (request.args.get('sort_dir') or 'asc').strip().lower()
    min_estrellas = 4
    query_ordenada, _, _ = _construir_query_clientes(
        buscar=buscar,
        solo_top=solo_top,
        sort_by=sort_by,
        sort_dir=sort_dir,
        min_estrellas=min_estrellas
    )
    clientes = query_ordenada.all()
    _inyectar_totales_clientes(clientes)

    timestamp = datetime.now()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Clientes'
    ws.merge_cells('A1:H1')
    ws['A1'] = 'Reporte de Clientes'
    ws['A1'].font = Font(bold=True, size=14, color='0F172A')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = f'Generado: {timestamp.strftime("%d/%m/%Y %H:%M")} | Orden: {sort_by} {sort_dir.upper()}'
    ws['A2'].font = Font(size=10, color='475569')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    ws.append([])
    header_row = 4
    ws.append([
        'Nombre',
        'RUC/CI',
        'Telefono',
        'Tipo',
        'Compras',
        'Reparaciones',
        'Total',
        'Nivel',
    ])
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='1D4ED8')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for cliente in clientes:
        ws.append([
            cliente.nombre or '',
            cliente.ruc_ci or '',
            cliente.telefono or '',
            cliente.tipo or '',
            int(cliente.total_compras or 0),
            int(cliente.total_reparaciones or 0),
            int(cliente.total_movimientos or 0),
            int(cliente.nivel_estrellas_seguro or 0),
        ])

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    primera_fila_datos = header_row + 1
    ultima_fila = ws.max_row
    for row in ws.iter_rows(min_row=header_row, max_row=ultima_fila, min_col=1, max_col=8):
        for cell in row:
            cell.border = thin_border
            if cell.row > header_row:
                cell.alignment = Alignment(vertical='center')
    for row in ws.iter_rows(min_row=primera_fila_datos, max_row=ultima_fila, min_col=5, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    fila_resumen = ultima_fila + 2
    ws[f'A{fila_resumen}'] = 'Total clientes'
    ws[f'A{fila_resumen}'].font = Font(bold=True, color='0F172A')
    ws[f'B{fila_resumen}'] = len(clientes)
    ws[f'B{fila_resumen}'].font = Font(bold=True, color='0F172A')
    ws[f'D{fila_resumen}'] = 'Movimientos acumulados'
    ws[f'D{fila_resumen}'].font = Font(bold=True, color='0F172A')
    ws[f'E{fila_resumen}'] = sum(int(cliente.total_movimientos or 0) for cliente in clientes)
    ws[f'E{fila_resumen}'].font = Font(bold=True, color='0F172A')

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:H{max(ws.max_row, 4)}'

    for col_idx in range(1, 9):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            text = str(cell.value)
            if len(text) > max_length:
                max_length = len(text)
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 45)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'clientes_{timestamp.strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@clientes_bp.route('/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo():
    """Crear nuevo cliente"""
    if not current_user.tiene_permiso('crear_cliente'):
        if getattr(current_user, 'modo_demo', False):
            flash('Modo demo: esta acción está deshabilitada.', 'warning')
        else:
            flash('No tienes permisos para crear clientes.', 'danger')
        return redirect(url_for('clientes.listar'))

    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        ruc_ci = request.form.get('ruc_ci', '').strip()
        telefono = request.form.get('telefono', '').strip()
        email = request.form.get('email', '').strip()
        direccion = request.form.get('direccion', '').strip()
        tipo = request.form.get('tipo', 'minorista')
        limite_credito = request.form.get('limite_credito', 0, type=float)
        nivel_estrellas = _parsear_nivel_estrellas(request.form.get('nivel_estrellas', 3))
        observacion = request.form.get('observacion', '').strip()
        
        if not nombre:
            flash('El nombre es obligatorio.', 'warning')
            return render_template(
                'clientes/form.html',
                cliente=None,
                prefill_nombre=nombre,
                prefill_telefono=telefono,
                observaciones=[]
            )
        
        cliente = Cliente(
            nombre=nombre,
            ruc_ci=ruc_ci,
            telefono=telefono,
            email=email,
            direccion=direccion,
            tipo=tipo,
            nivel_estrellas=nivel_estrellas,
            limite_credito=limite_credito,
            notas=observacion
        )
        
        db.session.add(cliente)
        db.session.commit()
        flash(f'Cliente "{nombre}" creado correctamente.', 'success')
        return redirect(url_for('clientes.listar'))
    
    prefill_nombre = request.args.get('nombre', '').strip()
    prefill_telefono = request.args.get('telefono', '').strip()
    return render_template(
        'clientes/form.html',
        cliente=None,
        prefill_nombre=prefill_nombre,
        prefill_telefono=prefill_telefono,
        observaciones=[]
    )


@clientes_bp.route('/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar(id):
    """Editar cliente"""
    if not current_user.tiene_permiso('editar_cliente'):
        if getattr(current_user, 'modo_demo', False):
            flash('Modo demo: esta acción está deshabilitada.', 'warning')
        else:
            flash('No tienes permisos para editar clientes.', 'danger')
        return redirect(url_for('clientes.listar'))

    cliente = Cliente.query.get_or_404(id)
    
    # No permitir editar Consumidor Final
    if cliente.id_cliente == 1:
        flash('No se puede editar el cliente Consumidor Final.', 'warning')
        return redirect(url_for('clientes.listar'))
    
    if request.method == 'POST':
        cliente.nombre = request.form.get('nombre', '').strip()
        cliente.ruc_ci = request.form.get('ruc_ci', '').strip()
        cliente.telefono = request.form.get('telefono', '').strip()
        cliente.email = request.form.get('email', '').strip()
        cliente.direccion = request.form.get('direccion', '').strip()
        cliente.tipo = request.form.get('tipo', 'minorista')
        cliente.nivel_estrellas = _parsear_nivel_estrellas(request.form.get('nivel_estrellas', 3))
        cliente.limite_credito = request.form.get('limite_credito', 0, type=float)
        cliente.notas = request.form.get('observacion', '').strip()
        
        db.session.commit()
        flash(f'Cliente "{cliente.nombre}" actualizado.', 'success')
        return redirect(url_for('clientes.listar'))
    
    observaciones = ClienteObservacion.query.filter_by(id_cliente=cliente.id_cliente)\
        .order_by(ClienteObservacion.fecha_observacion.desc())\
        .limit(100).all()
    return render_template('clientes/form.html', cliente=cliente, observaciones=observaciones)


@clientes_bp.route('/<int:id>/actualizar_estrellas', methods=['POST'])
@login_required
def actualizar_estrellas(id):
    es_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json
    next_url = (request.form.get('next') or '').strip()
    if not next_url.startswith('/'):
        next_url = ''
    destino = next_url or url_for('clientes.detalle', id=id)

    if not current_user.tiene_permiso('editar_cliente'):
        if getattr(current_user, 'modo_demo', False):
            if es_ajax:
                return jsonify({'success': False, 'error': 'Modo demo: esta acción está deshabilitada', 'modo_demo': True}), 403
            flash('Modo demo: esta acción está deshabilitada.', 'warning')
        else:
            if es_ajax:
                return jsonify({'success': False, 'error': 'No tienes permisos para editar clientes.', 'modo_demo': False}), 403
            flash('No tienes permisos para editar clientes.', 'danger')
        return redirect(destino)

    cliente = Cliente.query.get_or_404(id)

    if cliente.id_cliente == 1:
        if es_ajax:
            return jsonify({'success': False, 'error': 'No se puede editar el cliente Consumidor Final.'}), 400
        flash('No se puede editar el cliente Consumidor Final.', 'warning')
        return redirect(destino)

    cliente.nivel_estrellas = _parsear_nivel_estrellas(request.form.get('nivel_estrellas', 3))
    db.session.commit()
    if es_ajax:
        return jsonify({
            'success': True,
            'cliente': {
                'id_cliente': cliente.id_cliente,
                'nombre': cliente.nombre,
                'nivel_estrellas': cliente.nivel_estrellas_seguro
            }
        })
    flash(f'Estrellas de "{cliente.nombre}" actualizadas.', 'success')
    return redirect(destino)


@clientes_bp.route('/<int:id>/observaciones', methods=['POST'])
@login_required
def agregar_observacion(id):
    next_url = (request.form.get('next') or '').strip()
    if not next_url.startswith('/'):
        next_url = ''
    destino = next_url or url_for('clientes.detalle', id=id)

    if not current_user.tiene_permiso('editar_cliente'):
        if getattr(current_user, 'modo_demo', False):
            flash('Modo demo: esta acción está deshabilitada.', 'warning')
        else:
            flash('No tienes permisos para agregar observaciones.', 'danger')
        return redirect(destino)

    cliente = Cliente.query.get_or_404(id)
    texto = (request.form.get('observacion') or '').strip()
    if not texto:
        flash('Debes escribir una observación.', 'warning')
        return redirect(destino)

    observacion = ClienteObservacion(
        id_cliente=cliente.id_cliente,
        id_usuario=getattr(current_user, 'id_usuario', None),
        observacion=texto,
    )
    db.session.add(observacion)
    db.session.commit()
    flash('Observación registrada correctamente.', 'success')
    return redirect(destino)


@clientes_bp.route('/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar(id):
    """Eliminar (desactivar) cliente"""
    if not current_user.tiene_permiso('eliminar_cliente'):
        if getattr(current_user, 'modo_demo', False):
            flash('Modo demo: esta acción está deshabilitada.', 'warning')
        else:
            flash('No tienes permisos para eliminar clientes.', 'danger')
        return redirect(url_for('clientes.listar'))

    ok, _autorizacion = validar_autorizacion(
        request.form.get('id_autorizacion', type=int),
        'eliminar_cliente'
    )
    if not ok:
        flash('Se requiere autorización de administrador para eliminar clientes.', 'danger')
        return redirect(url_for('clientes.listar'))

    cliente = Cliente.query.get_or_404(id)
    
    if cliente.id_cliente == 1:
        flash('No se puede eliminar el cliente Consumidor Final.', 'danger')
        return redirect(url_for('clientes.listar'))
    
    cliente.activo = False
    db.session.commit()
    flash(f'Cliente "{cliente.nombre}" eliminado.', 'success')
    return redirect(url_for('clientes.listar'))


@clientes_bp.route('/buscar_json')
@login_required
def buscar_json():
    """Búsqueda de clientes en formato JSON"""
    if not (
        current_user.tiene_permiso('ver_clientes')
        or current_user.tiene_permiso('crear_reparacion')
        or current_user.tiene_permiso('editar_reparacion')
        or current_user.tiene_permiso('crear_venta')
        or current_user.tiene_permiso('ver_presupuestos_empresariales')
        or current_user.tiene_permiso('crear_presupuestos_empresariales')
    ):
        modo_demo = bool(getattr(current_user, 'modo_demo', False))
        payload = {'items': [], 'modo_demo': modo_demo}
        if modo_demo:
            payload['mensaje'] = 'Modo demo: esta acción está deshabilitada'
        else:
            payload['mensaje'] = 'Sin permisos para buscar clientes'
        return payload, 403

    q = request.args.get('q', '').strip()
    if q:
        clientes = Cliente.query.filter(
            Cliente.activo == True,
            db.or_(
                Cliente.nombre.ilike(f'%{q}%'),
                Cliente.ruc_ci.ilike(f'%{q}%')
            )
        ).limit(10).all()
    else:
        clientes = Cliente.query.filter(
            Cliente.activo == True
        ).order_by(Cliente.nombre.asc()).limit(10).all()
    
    return {
        'items': [{
            'id_cliente': c.id_cliente,
            'nombre': c.nombre,
            'ruc_ci': c.ruc_ci,
            'telefono': c.telefono,
            'direccion': c.direccion,
            'email': c.email,
            'tipo': c.tipo,
            'nivel_estrellas': c.nivel_estrellas_seguro,
            'observacion': c.notas or '',
            'limite_credito': float(c.limite_credito),
            'saldo_pendiente': float(c.saldo_pendiente)
        } for c in clientes]
    }


@clientes_bp.route('/buscar_por_telefono_json')
@login_required
def buscar_por_telefono_json():
    if not (
        current_user.tiene_permiso('ver_clientes')
        or current_user.tiene_permiso('crear_cliente')
        or current_user.tiene_permiso('editar_cliente')
        or current_user.tiene_permiso('crear_reparacion')
        or current_user.tiene_permiso('editar_reparacion')
        or current_user.tiene_permiso('crear_venta')
    ):
        modo_demo = bool(getattr(current_user, 'modo_demo', False))
        payload = {'success': False, 'cliente': None, 'modo_demo': modo_demo}
        if modo_demo:
            payload['error'] = 'Modo demo: esta acción está deshabilitada'
        else:
            payload['error'] = 'Sin permisos'
        return payload, 403

    telefono = request.args.get('telefono', '').strip()
    if not telefono:
        return {'success': True, 'cliente': None}

    tel_norm = normalizar_telefono(telefono) or telefono
    digits_input = re.sub(r'\D', '', telefono)
    digits_norm = re.sub(r'\D', '', tel_norm)

    condiciones = []
    if tel_norm:
        condiciones.append(Cliente.telefono == tel_norm)
    if telefono and telefono != tel_norm:
        condiciones.append(Cliente.telefono == telefono)
    if digits_norm:
        condiciones.append(Cliente.telefono == digits_norm)
        condiciones.append(Cliente.telefono.ilike(f'%{digits_norm}%'))
    elif digits_input:
        condiciones.append(Cliente.telefono.ilike(f'%{digits_input}%'))

    cliente = None
    if condiciones:
        cliente = Cliente.query.filter(
            Cliente.activo == True,
            db.or_(*condiciones)
        ).order_by(Cliente.id_cliente.desc()).first()

    if not cliente:
        return {'success': True, 'cliente': None}

    return {
        'success': True,
        'cliente': {
            'id_cliente': cliente.id_cliente,
            'nombre': cliente.nombre,
            'ruc_ci': cliente.ruc_ci,
            'telefono': cliente.telefono,
            'direccion': cliente.direccion,
            'email': cliente.email,
            'tipo': cliente.tipo,
            'nivel_estrellas': cliente.nivel_estrellas_seguro,
            'observacion': cliente.notas or '',
            'limite_credito': float(cliente.limite_credito or 0),
            'saldo_pendiente': float(cliente.saldo_pendiente or 0),
        }
    }


@clientes_bp.route('/crear_json', methods=['POST'])
@login_required
def crear_json():
    """Crear cliente desde JSON (para POS)"""
    if not current_user.tiene_permiso('crear_cliente'):
        if getattr(current_user, 'modo_demo', False):
            return {'success': False, 'error': 'Modo demo: esta acción está deshabilitada', 'modo_demo': True}, 403
        return {'success': False, 'error': 'Sin permisos', 'modo_demo': False}, 403

    data = request.get_json()
    
    nombre = data.get('nombre', '').strip()
    ruc_ci = data.get('ruc_ci', '').strip()
    
    if not nombre:
        return {'success': False, 'error': 'El nombre es obligatorio'}
        
    try:
        cliente = Cliente(
            nombre=nombre,
            ruc_ci=ruc_ci,
            telefono=data.get('telefono', '').strip(),
            direccion=data.get('direccion', '').strip(),
            email=data.get('email', '').strip(),
            tipo=data.get('tipo', 'minorista'),
            nivel_estrellas=_parsear_nivel_estrellas(data.get('nivel_estrellas', 3)),
            notas=(data.get('observacion') or data.get('notas') or '').strip(),
            limite_credito=data.get('limite_credito', 0)
        )
        
        db.session.add(cliente)
        db.session.commit()
        
        return {
            'success': True,
            'cliente': {
                'id_cliente': cliente.id_cliente,
                'nombre': cliente.nombre,
                'ruc_ci': cliente.ruc_ci,
                'telefono': cliente.telefono,
                'direccion': cliente.direccion,
                'email': cliente.email,
                'tipo': cliente.tipo,
                'nivel_estrellas': cliente.nivel_estrellas_seguro,
                'observacion': cliente.notas or '',
                'limite_credito': float(cliente.limite_credito)
            }
        }
    except Exception as e:
        db.session.rollback()
        return {'success': False, 'error': str(e)}


@clientes_bp.route('/<int:id>/limite_credito_json', methods=['POST'])
@login_required
def actualizar_limite_credito_json(id):
    """Actualiza solo el limite de credito para evitar pisar otros datos del cliente desde POS."""
    if not current_user.tiene_permiso('editar_cliente'):
        if getattr(current_user, 'modo_demo', False):
            return {'success': False, 'error': 'Modo demo: esta acción está deshabilitada', 'modo_demo': True}, 403
        return {'success': False, 'error': 'Sin permisos', 'modo_demo': False}, 403

    cliente = Cliente.query.get_or_404(id)
    if cliente.id_cliente == 1:
        return {'success': False, 'error': 'No se puede editar el Consumidor Final'}, 400

    data = request.get_json(silent=True) or {}
    try:
        limite_credito = max(0.0, float(data.get('limite_credito', 0) or 0))
    except (TypeError, ValueError):
        return {'success': False, 'error': 'El limite de credito es invalido'}, 400

    try:
        cliente.limite_credito = limite_credito
        db.session.commit()
        return {
            'success': True,
            'cliente': {
                'id_cliente': cliente.id_cliente,
                'limite_credito': float(cliente.limite_credito or 0),
                'saldo_pendiente': float(cliente.saldo_pendiente or 0),
                'credito_disponible': float(cliente.credito_disponible or 0),
            }
        }
    except Exception as e:
        db.session.rollback()
        return {'success': False, 'error': str(e)}, 500


@clientes_bp.route('/editar_json/<int:id>', methods=['POST'])
@login_required
def editar_json(id):
    """Editar cliente desde JSON (para POS)"""
    if not current_user.tiene_permiso('editar_cliente'):
        if getattr(current_user, 'modo_demo', False):
            return {'success': False, 'error': 'Modo demo: esta acción está deshabilitada', 'modo_demo': True}, 403
        return {'success': False, 'error': 'Sin permisos', 'modo_demo': False}, 403

    cliente = Cliente.query.get_or_404(id)
    
    if cliente.id_cliente == 1:
        return {'success': False, 'error': 'No se puede editar el Consumidor Final'}

    data = request.get_json()
    
    nombre = data.get('nombre', '').strip()
    if not nombre:
        return {'success': False, 'error': 'El nombre es obligatorio'}

    try:
        cliente.nombre = nombre
        cliente.ruc_ci = data.get('ruc_ci', '').strip()
        cliente.telefono = data.get('telefono', '').strip()
        cliente.direccion = data.get('direccion', '').strip()
        cliente.email = data.get('email', '').strip()
        cliente.tipo = data.get('tipo', 'minorista')
        cliente.nivel_estrellas = _parsear_nivel_estrellas(data.get('nivel_estrellas', 3))
        cliente.notas = (data.get('observacion') or data.get('notas') or '').strip()
        try:
            cliente.limite_credito = float(data.get('limite_credito', 0))
        except:
            cliente.limite_credito = 0
            
        db.session.commit()
        
        return {
            'success': True,
            'cliente': {
                'id_cliente': cliente.id_cliente,
                'nombre': cliente.nombre,
                'ruc_ci': cliente.ruc_ci,
                'telefono': cliente.telefono,
                'direccion': cliente.direccion,
                'email': cliente.email,
                'tipo': cliente.tipo,
                'nivel_estrellas': cliente.nivel_estrellas_seguro,
                'observacion': cliente.notas or '',
                'limite_credito': float(cliente.limite_credito)
            }
        }
    except Exception as e:
        db.session.rollback()
        return {'success': False, 'error': str(e)}


@clientes_bp.route('/<int:id>/historial_json')
@login_required
def historial_json(id):
    """Obtener historial de compras de un cliente"""
    if not (
        current_user.tiene_permiso('ver_clientes')
        or current_user.tiene_permiso('crear_venta')
    ):
        if getattr(current_user, 'modo_demo', False):
            return {'success': False, 'error': 'Modo demo: esta acción está deshabilitada', 'modo_demo': True}, 403
        return {'success': False, 'error': 'Sin permisos', 'modo_demo': False}, 403

    cliente = Cliente.query.get_or_404(id)
    
    # Obtener últimas 10 ventas completadas
    ventas = cliente.ventas.options(joinedload(Venta.cuenta_por_cobrar)).filter_by(estado='completada')\
        .order_by(db.desc('fecha_venta'))\
        .limit(10).all()
    reparaciones_raw = db.session.query(
        Reparacion.id_reparacion,
        Reparacion.fecha_ingreso,
        Reparacion.tipo_equipo,
        Reparacion.marca_modelo,
        Reparacion.estado,
        db.cast(Reparacion.costo_estimado, db.String).label('costo_estimado_raw'),
        db.cast(Reparacion.costo_final, db.String).label('costo_final_raw'),
    ).filter_by(cliente_id=id)\
        .order_by(Reparacion.fecha_ingreso.desc())\
        .limit(10).all()
        
    historial = []
    for v in ventas:
        # Resumen de items (ej: "Coca Cola x2, Pizza x1...")
        items = []
        for d in v.detalles.limit(3):
            items.append(f"{d.producto.nombre} x{d.cantidad}")
        
        if v.detalles.count() > 3:
            items.append(f"... (+{v.detalles.count() - 3})")
            
        historial.append({
            'id_venta': v.id_venta,
            'fecha': local_strftime(v.fecha_venta, '%d/%m/%Y %H:%M'),
            'total': float(v.total),
            'items_resumen': ", ".join(items),
            'numero_ticket': v.numero_comprobante or f"#{v.id_venta}",
            'tipo_venta': 'credito' if ((v.tipo_venta or 'contado').strip().lower() == 'credito') else 'contado',
            'saldo_pendiente': float((v.cuenta_por_cobrar.saldo_pendiente if v.cuenta_por_cobrar else v.saldo_pendiente) or 0),
        })
        
    # Estadísticas básicas
    total_gastado = db.session.query(db.func.sum(Venta.total))\
        .filter_by(id_cliente=id, estado='completada').scalar() or 0
        
    cantidad_compras = db.session.query(db.func.count(Venta.id_venta))\
        .filter_by(id_cliente=id, estado='completada').scalar() or 0

    historial_reparaciones = []
    total_reparaciones = 0.0
    for r in reparaciones_raw:
        monto_reparacion = _a_float_seguro(r.costo_final_raw)
        if monto_reparacion <= 0:
            monto_reparacion = _a_float_seguro(r.costo_estimado_raw)
        total_reparaciones += monto_reparacion
        historial_reparaciones.append({
            'id_reparacion': r.id_reparacion,
            'fecha': local_strftime(r.fecha_ingreso, '%d/%m/%Y %H:%M'),
            'equipo': f'{r.tipo_equipo} - {r.marca_modelo}',
            'estado': _estado_reparacion_display(r.estado),
            'monto': monto_reparacion
        })

    observaciones = ClienteObservacion.query.filter_by(id_cliente=id)\
        .order_by(ClienteObservacion.fecha_observacion.desc())\
        .limit(20).all()

    return {
        'success': True,
        'cliente': {
            'id_cliente': cliente.id_cliente,
            'nombre': cliente.nombre,
            'id': cliente.id_cliente,
            'ruc_ci': cliente.ruc_ci,
            'telefono': cliente.telefono,
            'tipo': cliente.tipo,
            'nivel_estrellas': cliente.nivel_estrellas_seguro,
            'observacion': cliente.notas or '',
            'url_historial': url_for('clientes.detalle', id=cliente.id_cliente)
        },
        'observaciones': [{
            'id_observacion': int(obs.id_observacion),
            'fecha': local_strftime(obs.fecha_observacion, '%d/%m/%Y %H:%M'),
            'observacion': obs.observacion or '',
            'usuario': (
                (obs.usuario.nombre_completo or '').strip()
                or (obs.usuario.username or '').strip()
            ) if obs.usuario else ''
        } for obs in observaciones],
        'historial': historial,
        'historial_reparaciones': historial_reparaciones,
        'estadisticas': {
            'total_gastado': float(total_gastado),
            'total_reparaciones': float(total_reparaciones),
            'gasto_total_general': float(total_gastado) + float(total_reparaciones),
            'cantidad_compras': cantidad_compras,
            'cantidad_reparaciones': len(historial_reparaciones),
            'promedio_compra': float(total_gastado / cantidad_compras) if cantidad_compras > 0 else 0
        }
    }


@clientes_bp.route('/<int:id>')
@login_required
def detalle(id):
    if not current_user.tiene_permiso('ver_clientes'):
        if getattr(current_user, 'modo_demo', False):
            flash('Modo demo: esta acción está deshabilitada.', 'warning')
        else:
            flash('No tienes permisos para ver clientes.', 'danger')
        return redirect(url_for('main.dashboard'))

    cliente = Cliente.query.get_or_404(id)
    cambios_fidelizacion = sincronizar_compras_fidelizacion_pendientes(id_cliente=cliente.id_cliente)
    cambios_fidelizacion += sincronizar_beneficios_vencidos(id_cliente=cliente.id_cliente, resumen_builder=beneficio_resumen_snapshot)
    if cambios_fidelizacion:
        db.session.commit()
        cliente = Cliente.query.get_or_404(id)
    compras_page = max(request.args.get('compras_page', 1, type=int), 1)
    reparaciones_page = max(request.args.get('reparaciones_page', 1, type=int), 1)
    historial_page = max(request.args.get('historial_page', 1, type=int), 1)
    compras_per_page = 10
    reparaciones_per_page = 10
    historial_per_page = 25

    ventas_pag = Venta.query.options(joinedload(Venta.cuenta_por_cobrar)).filter_by(id_cliente=id, estado='completada')\
        .order_by(Venta.fecha_venta.desc())\
        .paginate(page=compras_page, per_page=compras_per_page, error_out=False)
    ventas = ventas_pag.items

    reparaciones_base_q = db.session.query(
        Reparacion.id_reparacion,
        Reparacion.fecha_ingreso,
        Reparacion.tipo_equipo,
        Reparacion.marca_modelo,
        Reparacion.estado,
        db.cast(Reparacion.costo_estimado, db.String).label('costo_estimado_raw'),
        db.cast(Reparacion.costo_final, db.String).label('costo_final_raw'),
    ).filter_by(cliente_id=id)
    total_reparaciones_count = reparaciones_base_q.count()
    reparaciones_pages = max((total_reparaciones_count + reparaciones_per_page - 1) // reparaciones_per_page, 1)
    if reparaciones_page > reparaciones_pages:
        reparaciones_page = reparaciones_pages
    reparaciones_raw = reparaciones_base_q\
        .order_by(Reparacion.fecha_ingreso.desc())\
        .offset((reparaciones_page - 1) * reparaciones_per_page)\
        .limit(reparaciones_per_page).all()
    observaciones = ClienteObservacion.query.filter_by(id_cliente=id)\
        .order_by(ClienteObservacion.fecha_observacion.desc())\
        .limit(200).all()

    total_compras = db.session.query(db.func.sum(Venta.total))\
        .filter_by(id_cliente=id, estado='completada').scalar() or 0
    cantidad_compras = db.session.query(db.func.count(Venta.id_venta))\
        .filter_by(id_cliente=id, estado='completada').scalar() or 0
    ventas_historial = db.session.query(
        Venta.id_venta,
        Venta.fecha_venta,
        Venta.numero_comprobante,
        db.cast(Venta.total, db.String).label('total_raw'),
        Venta.tipo_venta,
        db.cast(Venta.saldo_pendiente, db.String).label('saldo_pendiente_raw'),
    ).filter_by(id_cliente=id, estado='completada').all()
    reparaciones_historial_raw = db.session.query(
        Reparacion.id_reparacion,
        Reparacion.fecha_ingreso,
        Reparacion.tipo_equipo,
        Reparacion.marca_modelo,
        Reparacion.estado,
        db.cast(Reparacion.costo_estimado, db.String).label('costo_estimado_raw'),
        db.cast(Reparacion.costo_final, db.String).label('costo_final_raw'),
    ).filter_by(cliente_id=id).all()

    total_reparaciones = 0.0
    historial = []
    reparaciones = []

    for venta in ventas_historial:
        monto = _a_float_seguro(venta.total_raw)
        saldo_pendiente = _a_float_seguro(getattr(venta, 'saldo_pendiente_raw', 0))
        tipo_venta = ((venta.tipo_venta or 'contado') if hasattr(venta, 'tipo_venta') else 'contado').strip().lower()
        estado_cobro = 'Pendiente' if saldo_pendiente > 0 else 'Pagada'
        detalle_partes = [
            venta.numero_comprobante or f'Venta #{venta.id_venta}',
            f'Tipo: {"Credito" if tipo_venta == "credito" else "Contado"}',
            estado_cobro,
        ]
        if saldo_pendiente > 0:
            detalle_partes.append(f'Saldo pendiente: Gs. {saldo_pendiente:,.0f}'.replace(',', '.'))
        historial.append({
            'tipo': 'compra',
            'id_referencia': int(venta.id_venta),
            'fecha': venta.fecha_venta,
            'fecha_texto': local_strftime(venta.fecha_venta, '%d/%m/%Y %H:%M'),
            'titulo': f'Compra #{venta.id_venta}',
            'detalle': ' | '.join(detalle_partes),
            'estado': estado_cobro,
            'monto': monto,
            'url': url_for('ventas.detalle', id=venta.id_venta)
        })

    for reparacion in reparaciones_raw:
        monto = _a_float_seguro(reparacion.costo_final_raw)
        if monto <= 0:
            monto = _a_float_seguro(reparacion.costo_estimado_raw)
        reparaciones.append({
            'id_reparacion': reparacion.id_reparacion,
            'fecha_ingreso': reparacion.fecha_ingreso,
            'tipo_equipo': reparacion.tipo_equipo,
            'marca_modelo': reparacion.marca_modelo,
            'estado': reparacion.estado,
            'estado_display': _estado_reparacion_display(reparacion.estado),
            'monto': monto,
        })

    for reparacion in reparaciones_historial_raw:
        monto = _a_float_seguro(reparacion.costo_final_raw)
        if monto <= 0:
            monto = _a_float_seguro(reparacion.costo_estimado_raw)
        total_reparaciones += monto
        historial.append({
            'tipo': 'reparacion',
            'id_referencia': int(reparacion.id_reparacion),
            'fecha': reparacion.fecha_ingreso,
            'fecha_texto': local_strftime(reparacion.fecha_ingreso, '%d/%m/%Y %H:%M'),
            'titulo': f'Reparación #{reparacion.id_reparacion}',
            'detalle': f'{reparacion.tipo_equipo} - {reparacion.marca_modelo}',
            'estado': _estado_reparacion_display(reparacion.estado),
            'monto': monto,
            'url': url_for('reparaciones.detalle', id=reparacion.id_reparacion)
        })

    historial.sort(key=_historial_sort_key, reverse=True)
    historial_total = len(historial)
    historial_pages = max((historial_total + historial_per_page - 1) // historial_per_page, 1)
    if historial_page > historial_pages:
        historial_page = historial_pages
    historial_inicio = (historial_page - 1) * historial_per_page
    historial_paginado = historial[historial_inicio:historial_inicio + historial_per_page]
    gasto_total_general = float(total_compras or 0) + float(total_reparaciones)

    return render_template(
        'clientes/detalle.html',
        cliente=cliente,
        observaciones=observaciones,
        ventas=ventas,
        reparaciones=reparaciones,
        historial=historial_paginado,
        compras_pag=ventas_pag,
        reparaciones_pag={
            'page': reparaciones_page,
            'pages': reparaciones_pages,
            'has_prev': reparaciones_page > 1,
            'has_next': reparaciones_page < reparaciones_pages,
            'prev_num': reparaciones_page - 1 if reparaciones_page > 1 else 1,
            'next_num': reparaciones_page + 1 if reparaciones_page < reparaciones_pages else reparaciones_pages,
            'per_page': reparaciones_per_page,
            'total': total_reparaciones_count,
        },
        historial_pag={
            'page': historial_page,
            'pages': historial_pages,
            'has_prev': historial_page > 1,
            'has_next': historial_page < historial_pages,
            'prev_num': historial_page - 1 if historial_page > 1 else 1,
            'next_num': historial_page + 1 if historial_page < historial_pages else historial_pages,
            'per_page': historial_per_page,
            'total': historial_total,
        },
        fidelizacion_resumen=obtener_resumen_beneficios_cliente(cliente.id_cliente),
        partial=request.args.get('partial'),
        estadisticas={
            'total_compras': float(total_compras or 0),
            'cantidad_compras': int(cantidad_compras or 0),
            'promedio_compra': float(total_compras / cantidad_compras) if cantidad_compras else 0,
            'total_reparaciones': float(total_reparaciones),
            'cantidad_reparaciones': int(total_reparaciones_count or 0),
            'gasto_total_general': gasto_total_general
        }
    )


@clientes_bp.route('/<int:id>/compras/<int:id_venta>/detalle_json')
@login_required
def detalle_compra_json(id, id_venta):
    if not current_user.tiene_permiso('ver_clientes'):
        if getattr(current_user, 'modo_demo', False):
            return jsonify({'success': False, 'error': 'Modo demo: esta acción está deshabilitada', 'modo_demo': True}), 403
        return jsonify({'success': False, 'error': 'Sin permisos', 'modo_demo': False}), 403

    cliente = Cliente.query.get_or_404(id)
    venta = Venta.query.filter_by(id_venta=id_venta, id_cliente=cliente.id_cliente, estado='completada').first_or_404()

    items = []
    for detalle in venta.detalles.all():
        nombre_producto = detalle.producto.nombre if detalle.producto else f'Producto #{detalle.id_producto}'
        items.append({
            'producto': nombre_producto,
            'cantidad': int(detalle.cantidad or 0),
            'precio_unitario': float(detalle.precio_unitario or 0),
            'subtotal': float(detalle.subtotal or 0),
        })

    pagos = []
    for pago in venta.pagos.all():
        pagos.append({
            'metodo': pago.metodo.nombre if pago.metodo else f'Método #{pago.id_metodo_pago}',
            'monto': float(pago.monto or 0),
        })

    return jsonify({
        'success': True,
        'tipo': 'compra',
        'id': int(venta.id_venta),
        'fecha': local_strftime(venta.fecha_venta, '%d/%m/%Y %H:%M'),
        'comprobante': venta.numero_comprobante or f'Venta #{venta.id_venta}',
        'total': float(venta.total or 0),
        'estado': 'Completada',
        'items': items,
        'pagos': pagos,
        'url': url_for('ventas.detalle', id=venta.id_venta),
    })


@clientes_bp.route('/<int:id>/reparaciones/<int:id_reparacion>/detalle_json')
@login_required
def detalle_reparacion_json(id, id_reparacion):
    if not current_user.tiene_permiso('ver_clientes'):
        if getattr(current_user, 'modo_demo', False):
            return jsonify({'success': False, 'error': 'Modo demo: esta acción está deshabilitada', 'modo_demo': True}), 403
        return jsonify({'success': False, 'error': 'Sin permisos', 'modo_demo': False}), 403

    cliente = Cliente.query.get_or_404(id)
    reparacion = db.session.query(
        Reparacion.id_reparacion,
        Reparacion.fecha_ingreso,
        Reparacion.tipo_equipo,
        Reparacion.marca_modelo,
        Reparacion.estado,
        Reparacion.falla_reportada,
        Reparacion.diagnostico_tecnico,
        Reparacion.solucion,
        Reparacion.accesorios,
        db.cast(Reparacion.costo_estimado, db.String).label('costo_estimado_raw'),
        db.cast(Reparacion.costo_final, db.String).label('costo_final_raw'),
    ).filter_by(
        id_reparacion=id_reparacion,
        cliente_id=cliente.id_cliente
    ).first_or_404()

    detalle_rows = db.session.query(
        DetalleReparacion.id_producto,
        DetalleReparacion.nombre_producto,
        DetalleReparacion.cantidad,
        db.cast(DetalleReparacion.precio_unitario, db.String).label('precio_unitario_raw'),
        db.cast(DetalleReparacion.subtotal, db.String).label('subtotal_raw'),
        DetalleReparacion.incluye_costo_final,
    ).filter_by(
        id_reparacion=reparacion.id_reparacion
    ).all()

    base_costo_final = _a_float_seguro(reparacion.costo_final_raw)
    extras_costo_final = sum(
        _a_float_seguro(detalle.subtotal_raw)
        for detalle in detalle_rows
        if bool(getattr(detalle, 'incluye_costo_final', False))
    )
    monto = base_costo_final + extras_costo_final
    if monto <= 0:
        monto = _a_float_seguro(reparacion.costo_estimado_raw)

    items = []
    for detalle in detalle_rows:
        items.append({
            'item': (detalle.nombre_producto or '').strip() or f'Ítem #{detalle.id_producto}',
            'cantidad': int(detalle.cantidad or 0),
            'precio_unitario': _a_float_seguro(detalle.precio_unitario_raw),
            'subtotal': _a_float_seguro(detalle.subtotal_raw),
            'incluye_costo_final': bool(getattr(detalle, 'incluye_costo_final', False)),
        })

    return jsonify({
        'success': True,
        'tipo': 'reparacion',
        'id': int(reparacion.id_reparacion),
        'fecha': local_strftime(reparacion.fecha_ingreso, '%d/%m/%Y %H:%M'),
        'equipo': f'{reparacion.tipo_equipo} - {reparacion.marca_modelo}',
        'estado': _estado_reparacion_display(reparacion.estado),
        'falla_reportada': (reparacion.falla_reportada or '').strip(),
        'diagnostico_tecnico': (reparacion.diagnostico_tecnico or '').strip(),
        'solucion': (reparacion.solucion or '').strip(),
        'accesorios': (reparacion.accesorios or '').strip(),
        'monto': monto,
        'items': items,
        'url': url_for('reparaciones.detalle', id=reparacion.id_reparacion),
    })
