import argparse
import os
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError

DEFAULT_EXCEL_NAME = "PRECIO_SISTEMA_NUEVO.xlsx"

COLUMN_ALIASES = {
    "codigo": "codigo",
    "categoria": "categoria_nombre",
    "nombre_del_producto": "nombre",
    "producto": "nombre",
    "descripcion": "descripcion",
    "marca": "marca",
    "modelo": "modelo",
    "color": "color",
    "capacidad": "capacidad",
    "precio_de_compra": "precio_compra",
    "precio_compra": "precio_compra",
    "precio_de_venta": "precio_venta",
    "precio_venta": "precio_venta",
    "precio_mayorista": "precio_mayorista",
    "iva": "porcentaje_iva",
    "stok": "stock_actual",
    "stock": "stock_actual",
    "stok_minio": "stock_minimo",
    "stock_minio": "stock_minimo",
    "stock_minimo": "stock_minimo",
}


@dataclass
class ImportStats:
    procesados: int = 0
    nuevos: int = 0
    actualizados: int = 0
    omitidos: int = 0
    errores: int = 0


def resolve_excel_file(file_path: str | None = None, base_dir: str | None = None) -> str:
    if file_path:
        return file_path

    base_dir = base_dir or os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, DEFAULT_EXCEL_NAME)


def normalizar_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("\ufeff", "")
    for char in ("/", "-", ".", "(", ")"):
        text = text.replace(char, " ")
    text = "_".join(text.split())
    return COLUMN_ALIASES.get(text, text)


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and float(value) == 0:
        return None
    text = str(value).strip()
    return text or None


def parse_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    text = str(value).strip()
    if not text:
        return Decimal("0.00")
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def parse_int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(parse_decimal(value))
    except Exception:
        return default


def row_to_payload(row: dict[str, Any]) -> dict[str, Any] | None:
    normalized = {normalizar_header(key): value for key, value in row.items()}
    codigo = clean_text(normalized.get("codigo"))
    nombre = clean_text(normalized.get("nombre"))

    if not codigo and not nombre:
        return None
    if not codigo:
        raise ValueError("Fila sin codigo")
    if not nombre:
        raise ValueError(f"Producto {codigo} sin nombre")

    return {
        "codigo": codigo,
        "categoria_nombre": clean_text(normalized.get("categoria_nombre")) or "Sin Categoria",
        "nombre": nombre,
        "descripcion": clean_text(normalized.get("descripcion")),
        "marca": clean_text(normalized.get("marca")),
        "modelo": clean_text(normalized.get("modelo")),
        "color": clean_text(normalized.get("color")),
        "capacidad": clean_text(normalized.get("capacidad")),
        "precio_compra": parse_decimal(normalized.get("precio_compra")),
        "precio_venta": parse_decimal(normalized.get("precio_venta")),
        "precio_mayorista": parse_decimal(normalized.get("precio_mayorista")),
        "porcentaje_iva": parse_int(normalized.get("porcentaje_iva"), default=10),
        "stock_actual": parse_int(normalized.get("stock_actual"), default=0),
        "stock_minimo": parse_int(normalized.get("stock_minimo"), default=5),
    }


def iter_excel_rows(file_path: str, sheet_name: str | None = None):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return

    for row in rows:
        if not row or all(value is None or str(value).strip() == "" for value in row):
            continue
        yield dict(zip(headers, row))


def get_or_create_categoria(nombre_categoria: str, Categoria, db, cache: dict[str, Any], dry_run: bool):
    nombre = (nombre_categoria or "Sin Categoria").strip()
    key = nombre.upper()
    if key in cache:
        return cache[key]

    categoria = Categoria.query.filter(db.func.upper(Categoria.nombre) == key).first()
    if categoria:
        cache[key] = categoria
        return categoria

    categoria = Categoria(nombre=nombre, descripcion="Importada desde Excel", activo=True)
    if not dry_run:
        db.session.add(categoria)
        db.session.flush()
    cache[key] = categoria
    return categoria


def upsert_producto(payload: dict[str, Any], Producto, Categoria, db, cache, dry_run: bool) -> str:
    categoria = get_or_create_categoria(payload["categoria_nombre"], Categoria, db, cache, dry_run)
    producto = Producto.query.filter_by(codigo=payload["codigo"]).first()
    values = {
        "nombre": payload["nombre"],
        "descripcion": payload["descripcion"],
        "id_categoria": categoria.id_categoria,
        "marca": payload["marca"],
        "modelo": payload["modelo"],
        "color": payload["color"],
        "capacidad": payload["capacidad"],
        "precio_compra": payload["precio_compra"],
        "precio_venta": payload["precio_venta"],
        "precio_mayorista": payload["precio_mayorista"],
        "porcentaje_iva": payload["porcentaje_iva"],
        "stock_actual": payload["stock_actual"],
        "stock_minimo": payload["stock_minimo"],
        "activo": True,
    }

    if producto:
        if not dry_run:
            for field, value in values.items():
                setattr(producto, field, value)
        return "actualizado"

    if not dry_run:
        db.session.add(Producto(codigo=payload["codigo"], **values))
    return "nuevo"


def importar_productos_excel(file_path: str, sheet_name: str | None = None, dry_run: bool = False, limit: int | None = None) -> ImportStats:
    from app import create_app, db
    from app.models.producto import Categoria, Producto

    stats = ImportStats()
    app = create_app()
    cache: dict[str, Any] = {}

    with app.app_context():
        for raw_row in iter_excel_rows(file_path, sheet_name):
            if limit is not None and stats.procesados >= limit:
                break
            try:
                payload = row_to_payload(raw_row)
                if payload is None:
                    stats.omitidos += 1
                    continue
                result = upsert_producto(payload, Producto, Categoria, db, cache, dry_run)
                if result == "nuevo":
                    stats.nuevos += 1
                else:
                    stats.actualizados += 1
                stats.procesados += 1
            except (IntegrityError, Exception) as exc:
                db.session.rollback()
                stats.errores += 1
                print(f"Error en fila: {exc}")

        if dry_run:
            db.session.rollback()
        else:
            db.session.commit()
    return stats


def main():
    parser = argparse.ArgumentParser(description="Importar productos desde Excel")
    parser.add_argument("--file", default=None, help="Ruta del archivo .xlsx")
    parser.add_argument("--sheet", default=None, help="Nombre de hoja a importar")
    parser.add_argument("--dry-run", action="store_true", help="Simular sin guardar cambios")
    parser.add_argument("--limit", type=int, default=None, help="Limitar cantidad de filas procesadas")
    args = parser.parse_args()

    file_path = resolve_excel_file(args.file)
    if not os.path.exists(file_path):
        raise SystemExit(f"No existe el archivo: {file_path}")

    stats = importar_productos_excel(file_path, args.sheet, args.dry_run, args.limit)
    print("\nResumen de importacion")
    print(f"Procesados:    {stats.procesados}")
    print(f"Nuevos:        {stats.nuevos}")
    print(f"Actualizados:  {stats.actualizados}")
    print(f"Omitidos:      {stats.omitidos}")
    print(f"Errores:       {stats.errores}")
    if args.dry_run:
        print("Modo dry-run: no se guardaron cambios.")


if __name__ == "__main__":
    main()
